import asyncio
import json
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, StateFilter
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile, Document
from pathlib import Path
from openpyxl import load_workbook, Workbook
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram import Router
from sqlalchemy import insert, select, update, delete
from sqlalchemy.ext.asyncio import AsyncSession
from shared.config import get_settings
from shared.db import SessionLocal, init_db
from db import models

settings = get_settings()
bot = Bot(settings.BOT1_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()
router = Router()
dp.include_router(router)

# Папка с правилами/договорами
RULES_DIR = Path(__file__).resolve().parent / "rules"
BASE_PRICE_RUB = 2000

async def get_discount_percent(session: AsyncSession, user_identifier: str | None) -> float:
    if not user_identifier:
        return 0.0
    res = await session.execute(
        select(models.Discount.percent).where(models.Discount.user_identifier == user_identifier)
    )
    row = res.first()
    return float(row[0]) if row else 0.0

async def calc_price_with_discount_db(session: AsyncSession, username_with_at: str | None) -> int:
    percent = await get_discount_percent(session, username_with_at)
    price = round(BASE_PRICE_RUB * (1.0 - percent / 100.0))
    if price < 0:
        price = 0
    return int(price)


# ================== Импорт скидок из Excel в отдельном чате ==================

# Импорт разрешён: либо из чата DISCOUNT_CHAT_ID, либо от ADMIN_CHAT_ID (например, в ЛС)
@router.message(F.document, (F.chat.id == settings.DISCOUNT_CHAT_ID) | (F.from_user.id == settings.ADMIN_CHAT_ID))
async def import_discounts_from_excel(m: Message):
    doc: Document = m.document  # type: ignore
    if not doc.file_name.lower().endswith((".xlsx", ".xlsm")):
        await m.answer("Пожалуйста, отправьте Excel-файл формата .xlsx/.xlsm")
        return
    await m.answer("Файл принят, начинаю обработку…", reply_markup=kb(["Сбросить скидки", "Help"], row=2))
    try:
        print(f"[DISCOUNTS] Received file: name={doc.file_name}, size={doc.file_size}, mime={doc.mime_type}")
    except Exception:
        pass

    file = await bot.get_file(doc.file_id)
    file_path = file.file_path
    # Скачиваем файл во временную директорию
    tmp_dir = Path("./data/tmp")
    tmp_dir.mkdir(parents=True, exist_ok=True)
    local_path = tmp_dir / doc.file_name
    print(f"[DISCOUNTS] Downloading from TG path: {file_path} -> {local_path}")
    await bot.download_file(file_path, destination=str(local_path))
    print(f"[DISCOUNTS] File downloaded: {local_path.exists()} size={local_path.stat().st_size if local_path.exists() else 'n/a'}")

    try:
        print("[DISCOUNTS] Opening workbook…")
        wb = load_workbook(filename=str(local_path), read_only=True, data_only=True)
        ws = wb.active
        print(f"[DISCOUNTS] Active sheet: {ws.title}")
        updates = 0
        inserts = 0
        skipped = 0
        summary_rows: list[tuple[str, float | None, float | None, str]] = []  # (user, old, new, action)
        async with SessionLocal() as session:  # type: AsyncSession
            row_index = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                row_index += 1
                if not row:
                    continue
                user_ident = (str(row[0]).strip() if row[0] is not None else "")
                if not user_ident:
                    continue
                # Нормализуем к виду @username
                if not user_ident.startswith("@"):
                    user_ident = f"@{user_ident}"
                # Парсим процент
                percent_raw = row[1]
                try:
                    if isinstance(percent_raw, str) and percent_raw.endswith("%"):
                        percent_val = float(percent_raw[:-1])
                    else:
                        percent_val = float(percent_raw)
                except Exception:
                    skipped += 1
                    print(f"[DISCOUNTS] Row {row_index}: failed to parse percent '{row[1]}' -> skipped")
                    continue

                res = await session.execute(
                    select(models.Discount).where(models.Discount.user_identifier == user_ident)
                )
                existing = res.scalar_one_or_none()
                if existing is None:
                    await session.execute(
                        insert(models.Discount).values(user_identifier=user_ident, percent=percent_val)
                    )
                    inserts += 1
                    print(f"[DISCOUNTS] Row {row_index}: add {user_ident} -> {percent_val}%")
                    summary_rows.append((user_ident, None, percent_val, "added"))
                else:
                    old_percent = float(existing.percent or 0.0)
                    if abs(old_percent - percent_val) > 1e-9:
                        await session.execute(
                            update(models.Discount)
                            .where(models.Discount.id == existing.id)
                            .values(percent=percent_val)
                        )
                        updates += 1
                        print(f"[DISCOUNTS] Row {row_index}: update {user_ident} {old_percent}% -> {percent_val}%")
                        summary_rows.append((user_ident, old_percent, percent_val, "updated"))
                    else:
                        skipped += 1
                        print(f"[DISCOUNTS] Row {row_index}: unchanged {user_ident} {old_percent}%")
                        summary_rows.append((user_ident, old_percent, old_percent, "unchanged"))
            await session.commit()
        print(f"[DISCOUNTS] Import done: inserts={inserts}, updates={updates}, unchanged={skipped}")
        # Формируем сводную таблицу Excel
        summary_path = tmp_dir / (local_path.stem + "_summary.xlsx")
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Итоги импорта"
        ws_out.append(["Пользователь", "Старый %", "Новый %", "Действие"])
        for user, oldp, newp, action in summary_rows:
            ws_out.append([user, oldp if oldp is not None else "—", newp if newp is not None else "—", action])
        wb_out.save(str(summary_path))

        await m.answer(
            f"Импорт завершен. Добавлено: {inserts}, обновлено: {updates}, без изменений: {skipped}.",
            reply_markup=kb(["Сбросить скидки", "Help"], row=2)
        )
        await m.answer_document(
            FSInputFile(str(summary_path)),
            caption="Сводная таблица по импорту скидок"
        )
    except Exception as e:
        print(f"[DISCOUNTS] Import error: {e}")
        await m.answer(f"Ошибка импорта: {e}", reply_markup=kb(["Сбросить скидки", "Help"], row=2))
    finally:
        try:
            local_path.unlink(missing_ok=True)
        except Exception:
            pass
        try:
            summary_path.unlink(missing_ok=True)  # type: ignore[name-defined]
        except Exception:
            pass

# Сообщение для любых других документов не из разрешённого чата
@router.message(F.document)
async def reject_import_from_other_chats(m: Message):
    try:
        await m.answer(
            f"Импорт скидок разрешён только в специальном чате. Текущий chat_id: {m.chat.id}.\n"
            f"Ожидаемый DISCOUNT_CHAT_ID: {settings.DISCOUNT_CHAT_ID}."
        )
    except Exception:
        pass

# ================== Сброс всех скидок в DISCOUNT_CHAT ==================
@router.message(
    (F.chat.id == settings.DISCOUNT_CHAT_ID) & (F.text.lower() == "/discount_reset")
    |
    ((F.from_user.id == settings.ADMIN_CHAT_ID) & (F.text.lower() == "/discount_reset"))
)
async def reset_all_discounts(m: Message):
    try:
        await m.answer("Подтверждаю команду. Начинаю сброс всех скидок…", reply_markup=kb(["Сбросить скидки", "Help"], row=2))
        print("[DISCOUNTS] Reset command received. Deleting all discounts…")
        async with SessionLocal() as session:  # type: AsyncSession
            await session.execute(delete(models.Discount))
            await session.commit()
        await m.answer("Все скидки удалены.", reply_markup=kb(["Сбросить скидки", "Help"], row=2))
        print("[DISCOUNTS] All discounts deleted.")
    except Exception as e:
        print(f"[DISCOUNTS] Reset error: {e}")
        await m.answer(f"Ошибка при удалении скидок: {e}")

# Меню для DISCOUNT_CHAT: показать клавиатуру и Help
@router.message(F.chat.id == settings.DISCOUNT_CHAT_ID, (F.text.lower() == "/start") | (F.text.lower() == "help") | (F.text.lower() == "/help") | (F.text.lower() == "меню"))
async def discount_chat_help(m: Message):
    help_text = (
        "Управление скидками:\n"
        "- Отправьте Excel (.xlsx/.xlsm) с колонками: A=@username, B=скидка (напр. 10 или 10%).\n"
        "- Кнопка ‘Сбросить скидки’ удалит все записи.\n"
        "- Кнопка ‘Help’ — это сообщение.\n"
        f"DISCOUNT_CHAT_ID: {settings.DISCOUNT_CHAT_ID}\n"
    )
    await m.answer(help_text, reply_markup=kb(["Сбросить скидки", "Help"], row=2))

# Альтернативная кнопка без слеша в DISCOUNT_CHAT
@router.message(F.chat.id == settings.DISCOUNT_CHAT_ID, F.text == "Сбросить скидки")
async def discount_chat_reset_button(m: Message):
    # переиспользуем команду
    await reset_all_discounts(m)

# Показать клавиатуру в DISCOUNT_CHAT для любых других текстовых сообщений
@router.message(
    (F.chat.id == settings.DISCOUNT_CHAT_ID)
    & F.text
    & (F.text.lower() != "/start")
    & (F.text.lower() != "/help")
    & (F.text.lower() != "help")
    & (F.text.lower() != "меню")
    & (F.text.lower() != "/discount_reset")
    & (F.text != "Сбросить скидки")
)
async def discount_chat_always_keyboard(m: Message):
    await m.answer("Меню скидок:", reply_markup=kb(["Сбросить скидки", "Help"], row=2))

def kb(options: list[str], one_time: bool = False, row: int = 2) -> ReplyKeyboardMarkup:
    rows = []
    for i in range(0, len(options), row):
        rows.append([KeyboardButton(text=o) for o in options[i:i+row]])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True, one_time_keyboard=one_time)

LEVELS = ["A1", "A2", "B1", "B2", "C1", "C2"]
ACK_BTN = ["Ознакомлен(а)"]
YES_NO = ["Да", "Нет"]
TIMEZONES = [
    "Калининградское время (UTC+2)",
    "Московское время (UTC+3)",
    "Самарское время (UTC+4)",
    "Екатеринбургское время (UTC+5)",
    "Омское время (UTC+6)",
    "Красноярское время (UTC+7)",
    "Иркутское время (UTC+8)",
    "Якутское время (UTC+9)",
    "Владивостокское время (UTC+10)",
    "Магаданское время (UTC+11)",
    "Камчатское время (UTC+12)",
    "Другой",
]
PRIORITY = ["Улучшить разговорный английский", "Преодолеть языковой барьер", "Поддерживать регулярную практику", "Найти единомышленника", "Другое"]
DAYS = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота"]
TIME_OF_DAY = ["Утром","Днем","Вечером","Не имеет значения"]
STRICTNESS = ["Очень важно","Желательно, но могу подстраиваться","Готов(а) к разным вариантам"]
INITIATIVE = ["Люблю предлагать сам(а)","Предпочитаю, когда предлагают другие","Открыт(а) к обоим вариантам"]
GENDER = ["Девушка","Парень","Не имеет значения"]

class FormStates(StatesGroup):
    FULL_NAME = State()
    AGE = State()
    TIMEZONE = State()
    CITY = State()
    TG_CONTACT_CHOICE = State()
    TG_USERNAME = State()
    ALT_CONTACT = State()
    LEVEL = State()
    PARTICIPATED = State()
    PRIORITY = State()
    GOALS = State()
    HOBBIES = State()
    TOPICS = State()
    DAYS = State()
    TIME_OF_DAY = State()
    STRICTNESS = State()
    INITIATIVE = State()
    PARTNER_GENDER = State()
    OTHER_NOTES = State()
    ACCEPT_RULES = State()
    ACCEPT_OFFER = State()
    ACCEPT_NDA = State()
    PAYMENT = State()

@router.message(CommandStart())
async def start(m: Message, state: FSMContext):
    # Запрет на повторное заполнение анкеты
    async with SessionLocal() as session:  # type: AsyncSession
        exists_res = await session.execute(
            select(models.Form.id).where(models.Form.tg_user_id == m.from_user.id)
        )
        if exists_res.first():
            await m.answer("Вы уже заполнили анкету. Повторная отправка недоступна.", reply_markup=kb(["Другое"], row=1))
            return

    await state.clear()
    # Приветственное сообщение + кнопки
    welcome_text = (
        "Привет! Добро пожаловать в проект Talko.\n"
        "Я помогу тебе записаться на программу Study Buddy и расскажу все подробности\n"
        "Если хочешь записаться на программу - нажми кнопку \"Записаться\""
    )
    await m.answer(welcome_text, reply_markup=kb(["Записаться", "Другое"], one_time=False, row=2))
    await m.answer("По всем вопросом обращайтесь к @polyaa_makarova")

@router.message(F.text == "Записаться")
async def on_enroll(m: Message, state: FSMContext):
    # Проверка на повторное заполнение при нажатии "Записаться"
    async with SessionLocal() as session:  # type: AsyncSession
        exists_res = await session.execute(
            select(models.Form.id).where(models.Form.tg_user_id == m.from_user.id)
        )
        if exists_res.first():
            await m.answer("Вы уже заполнили анкету. Повторная отправка недоступна.", reply_markup=kb(["Другое"], row=1))
            return
    await m.answer("Заполним анкету.\nВведите, пожалуйста, ФИО:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.FULL_NAME)

@router.message(StateFilter(None), F.text == "Другое")
async def on_other(m: Message):
    await m.answer("Выберите опцию:", reply_markup=kb(["Поддержка", "Сообщество", "Назад"], one_time=False, row=2))

@router.message(StateFilter(None), F.text == "Назад")
async def on_back(m: Message):
    await m.answer("Главное меню:", reply_markup=kb(["Записаться", "Другое"], one_time=False, row=2))

@router.message(StateFilter(None), F.text == "Поддержка")
async def on_support(m: Message):
    await m.answer("По всем вопросом обращайтесь к @polyaa_makarova")

@router.message(StateFilter(None), F.text == "Сообщество")
async def on_community(m: Message):
    await m.answer("Подписывайтесь на наш канал https://t.me/Talko_1")

@router.message(FormStates.FULL_NAME)
async def full_name(m: Message, state: FSMContext):
    await state.update_data(full_name=m.text.strip())
    await m.answer("Возраст (14+):", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.AGE)

@router.message(FormStates.AGE)
async def age(m: Message, state: FSMContext):
    text = (m.text or "").strip()
    if not text.isdigit():
        await m.answer("Пожалуйста, введите возраст цифрами (например, 23).")
        return
    age_num = int(text)
    if age_num < 14 or age_num > 120:
        await m.answer("Пожалуйста, введите реальный возраст от 14 до 120.")
        return
    await state.update_data(age=str(age_num))
    await m.answer("Выберите ваш часовой пояс:", reply_markup=kb(TIMEZONES, row=2))
    await state.set_state(FormStates.TIMEZONE)

@router.message(FormStates.TIMEZONE, F.text.in_(TIMEZONES))
async def tz(m: Message, state: FSMContext):
    if m.text == "Другой":
        await m.answer("Введите смещение относительно UTC в формате +N или -N (например, +3):", reply_markup=ReplyKeyboardRemove())
        # временно используем то же состояние, обработаем свободный ввод ниже
        return
    await state.update_data(timezone=m.text.strip())
    suggested = (m.from_user.username or "").strip()
    if suggested:
        await state.update_data(tg_username=f"@{suggested}")
    await m.answer("Город проживания:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.CITY)

@router.message(FormStates.TIMEZONE)
async def tz_manual(m: Message, state: FSMContext):
    text = (m.text or "").strip().replace(" ", "")
    if not text:
        await m.answer("Введите смещение, например +3 или -2.")
        return
    sign = 1
    if text[0] in ["+", "-"]:
        sign = 1 if text[0] == "+" else -1
        num_part = text[1:]
    else:
        num_part = text
    if not num_part.isdigit():
        await m.answer("Пожалуйста, используйте только цифры после знака (например, +3).")
        return
    hours = int(num_part) * sign
    if hours < -12 or hours > 14:
        await m.answer("Смещение должно быть в диапазоне от -12 до +14.")
        return
    await state.update_data(timezone=f"UTC{('+' if hours>=0 else '')}{hours}")
    suggested = (m.from_user.username or "").strip()
    if suggested:
        await state.update_data(tg_username=f"@{suggested}")
    await m.answer("Город проживания:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.CITY)

@router.message(FormStates.CITY)
async def city(m: Message, state: FSMContext):
    await state.update_data(city=m.text.strip())
    await m.answer("Оставьте контакт для связи", reply_markup=kb(["Мой Telegram", "Другой контакт"], one_time=True))
    await state.set_state(FormStates.TG_CONTACT_CHOICE)

@router.message(FormStates.TG_CONTACT_CHOICE, F.text.in_(["Мой Telegram", "Другой контакт"]))
async def tg_contact_choice(m: Message, state: FSMContext):
    if m.text == "Мой Telegram":
        existing = (m.from_user.username or "").strip()
        if not existing:
            await m.answer("У вас нет публичного @username. Введите контакт вручную (например, номер телефона, почту или @ник):", reply_markup=ReplyKeyboardRemove())
            await state.set_state(FormStates.TG_USERNAME)
        else:
            await state.update_data(tg_username=f"@{existing}")
            await m.answer("Уровень языка (если не знаешь свой уровень - пройди тест: https://www.cambridgeenglish.org/test-your-english/)", reply_markup=kb(LEVELS))
            await state.set_state(FormStates.LEVEL)
    else:
        await m.answer("Введите контакт (почта/телефон/@ник):", reply_markup=ReplyKeyboardRemove())
        await state.set_state(FormStates.ALT_CONTACT)

@router.message(FormStates.ALT_CONTACT)
async def alt_contact(m: Message, state: FSMContext):
    await state.update_data(alt_contact=m.text.strip())
    await m.answer("Уровень языка (если не знаешь свой уровень - пройди тест: https://www.cambridgeenglish.org/test-your-english/)", reply_markup=kb(LEVELS))
    await state.set_state(FormStates.LEVEL)

@router.message(FormStates.TG_USERNAME)
async def uname(m: Message, state: FSMContext):
    await state.update_data(tg_username=m.text.strip())
    await m.answer("Уровень языка (если не знаешь свой уровень - пройди тест: https://www.cambridgeenglish.org/test-your-english/)", reply_markup=kb(LEVELS))
    await state.set_state(FormStates.LEVEL)

@router.message(FormStates.LEVEL, F.text.in_(LEVELS))
async def level(m: Message, state: FSMContext):
    await state.update_data(level=m.text)
    await m.answer("Участвовал(а) ли ранее в языковых клубах или парной практике?", reply_markup=kb(YES_NO))
    await state.set_state(FormStates.PARTICIPATED)

@router.message(FormStates.PARTICIPATED, F.text.in_(YES_NO))
async def participated(m: Message, state: FSMContext):
    await state.update_data(participated_before=m.text)
    await m.answer("Что для тебя важнее всего в парной практике?", reply_markup=kb(PRIORITY, one_time=True))
    await state.set_state(FormStates.PRIORITY)

@router.message(FormStates.PRIORITY, F.text.in_(PRIORITY))
async def priority(m: Message, state: FSMContext):
    await state.update_data(priority=m.text)
    await m.answer("Для каких целей учишь английский? (например, ЕГЭ, олимпиады или хобби)", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.GOALS)

@router.message(FormStates.GOALS)
async def goals(m: Message, state: FSMContext):
    await state.update_data(goals=m.text.strip())
    await m.answer("Напиши о своих хобби и интересах:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.HOBBIES)

@router.message(FormStates.HOBBIES)
async def hobbies(m: Message, state: FSMContext):
    await state.update_data(hobbies=m.text.strip())
    await m.answer("Какие темы для обсуждений тебе особенно интересны?", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.TOPICS)

@router.message(FormStates.TOPICS)
async def topics(m: Message, state: FSMContext):
    await state.update_data(topics=m.text.strip(), days_selected=[])
    await m.answer("Выбери дни недели (нажимайте на нужный день недели, когда выберете все удобные дни, нажмите \"Готово\", если ошиблись при выборе дня, нажмите \"Сбросить\")", reply_markup=kb(DAYS + ["Готово","Сбросить"], row=3))
    await state.set_state(FormStates.DAYS)

@router.message(FormStates.DAYS, F.text == "Сбросить")
async def days_reset(m: Message, state: FSMContext):
    await state.update_data(days_selected=[])
    await m.answer("Сброшено. Выбирай дни, затем 'Готово'.")

@router.message(FormStates.DAYS, F.text == "Готово")
async def days_done(m: Message, state: FSMContext):
    data = await state.get_data()
    selected = data.get("days_selected", [])
    if not selected:
        await m.answer("Пожалуйста, выбери хотя бы один день.")
        return
    await state.update_data(days=selected)
    await m.answer("Тебе удобнее заниматься утром, днем или вечером?", reply_markup=kb(TIME_OF_DAY, one_time=True))
    await state.set_state(FormStates.TIME_OF_DAY)

@router.message(FormStates.DAYS, F.text.in_(DAYS))
async def days_add(m: Message, state: FSMContext):
    data = await state.get_data()
    selected = set(data.get("days_selected", []))
    if m.text in selected:
        selected.remove(m.text)
    else:
        selected.add(m.text)
    await state.update_data(days_selected=list(selected))
    await m.answer(f"Выбрано: {', '.join(selected) or 'пока ничего'}")

@router.message(FormStates.TIME_OF_DAY, F.text.in_(TIME_OF_DAY))
async def time_of_day(m: Message, state: FSMContext):
    await state.update_data(time_of_day=m.text)
    await m.answer("Насколько важно соблюдение расписания?", reply_markup=kb(STRICTNESS, one_time=True))
    await state.set_state(FormStates.STRICTNESS)

@router.message(FormStates.STRICTNESS, F.text.in_(STRICTNESS))
async def strictness(m: Message, state: FSMContext):
    await state.update_data(schedule_strictness=m.text)
    await m.answer("Как ты относишься к самостоятельной инициативе в паре (например, предложить дополнительную активность)", reply_markup=kb(INITIATIVE, one_time=True))
    await state.set_state(FormStates.INITIATIVE)

@router.message(FormStates.INITIATIVE, F.text.in_(INITIATIVE))
async def initiative(m: Message, state: FSMContext):
    await state.update_data(initiative=m.text)
    await m.answer("С кем бы ты хотел(а) быть в паре?", reply_markup=kb(GENDER, one_time=True))
    await state.set_state(FormStates.PARTNER_GENDER)

@router.message(FormStates.PARTNER_GENDER, F.text.in_(GENDER))
async def partner_gender(m: Message, state: FSMContext):
    await state.update_data(preferred_partner_gender=m.text)
    await m.answer("Есть ли что-то важное для подбора? (введи текст или '-' если нет)", reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.OTHER_NOTES)

@router.message(FormStates.OTHER_NOTES)
async def other_notes(m: Message, state: FSMContext):
    await state.update_data(other_notes=m.text.strip())
    rules_path = RULES_DIR / "Пакт правил Talko.pdf"
    await m.answer_document(
        FSInputFile(str(rules_path)),
        caption="Подтверждаешь ознакомление с Пактом правил Talko?",
        reply_markup=kb(ACK_BTN, one_time=True)
    )
    await state.set_state(FormStates.ACCEPT_RULES)

@router.message(FormStates.ACCEPT_RULES, F.text.in_(ACK_BTN))
async def accept_rules(m: Message, state: FSMContext):
    await state.update_data(accepted_rules=True)
    offer_path = RULES_DIR / "Договор оферты Talko.pdf"
    await m.answer_document(
        FSInputFile(str(offer_path)),
        caption="Подтверждаешь ознакомление с Договором оферты Talko?",
        reply_markup=kb(ACK_BTN, one_time=True)
    )
    await state.set_state(FormStates.ACCEPT_OFFER)

@router.message(FormStates.ACCEPT_OFFER, F.text.in_(ACK_BTN))
async def accept_offer(m: Message, state: FSMContext):
    await state.update_data(accepted_offer=True)
    nda_path = RULES_DIR / "Договор NDA Talko.pdf"
    await m.answer_document(
        FSInputFile(str(nda_path)),
        caption="Подтверждаешь ознакомление с Договором NDA Talko?",
        reply_markup=kb(ACK_BTN, one_time=True)
    )
    await state.set_state(FormStates.ACCEPT_NDA)

@router.message(FormStates.ACCEPT_NDA, F.text.in_(ACK_BTN))
async def accept_nda(m: Message, state: FSMContext):
    await state.update_data(accepted_nda=True)
    # Рассчитываем цену с учетом скидки по @username
    username = (m.from_user.username or "").strip()
    username_with_at = f"@{username}" if username else None
    async with SessionLocal() as session:  # type: AsyncSession
        price = await calc_price_with_discount_db(session, username_with_at)
    if price < BASE_PRICE_RUB:
        payment_text = (
            f"Требуется оплатить доступ к участию в проекте. Переведите {price} руб. на следующие реквизиты:\n"
            "4276400093340025\n"
            "Полина Павловна М.\n\n"
            "Отправьте скриншот с переводом в этот чат"
        )
    else:
        payment_text = (
            f"Требуется оплатить доступ к участию в проекте. Переведите {BASE_PRICE_RUB} руб. на следующие реквизиты:\n"
            "4276400093340025\n"
            "Полина Павловна М.\n\n"
            "Отправьте скриншот с переводом в этот чат"
        )
    await m.answer(payment_text, reply_markup=ReplyKeyboardRemove())
    await state.set_state(FormStates.PAYMENT)

@router.message(FormStates.PAYMENT, F.photo | F.document["mime_type"].in_(["image/jpeg","image/png","image/webp"]))
async def payment_photo(m: Message, state: FSMContext):
    file_id = None
    if m.photo:
        file_id = m.photo[-1].file_id
    elif m.document:
        file_id = m.document.file_id
    await state.update_data(repost_photo_file_id=file_id)
    data = await state.get_data()

    to_save = {
        "tg_user_id": m.from_user.id,
        "full_name": data["full_name"],
        "age": data["age"],
        "timezone": data["timezone"],
        "city": (data.get("city") or ""),
        "tg_username": data.get("tg_username", ""),
        "alt_contact": data.get("alt_contact", ""),
        "level": data["level"],
        "participated_before": data["participated_before"],
        "priority": data["priority"],
        "hobbies": data["hobbies"],
        "goals": (data.get("goals") or ""),
        "topics": data["topics"],
        "days": json.dumps(data["days"]),
        "time_of_day": data["time_of_day"],
        "schedule_strictness": data["schedule_strictness"],
        "initiative": data["initiative"],
        "preferred_partner_gender": data["preferred_partner_gender"],
        "other_notes": data.get("other_notes", ""),
        "accepted_rules": bool(data["accepted_rules"]),
        "accepted_offer": bool(data["accepted_offer"]),
        "accepted_nda": bool(data["accepted_nda"]),
        "repost_photo_file_id": file_id,
    }

    async with SessionLocal() as session:  # type: AsyncSession
        # Дополнительная проверка на гонки: не вставлять дубликат, если уже есть запись
        exists_res = await session.execute(
            select(models.Form.id).where(models.Form.tg_user_id == m.from_user.id)
        )
        if exists_res.first():
            await m.answer("Анкета уже существует. Повторная отправка недоступна.")
            await state.clear()
            return

        await session.execute(insert(models.Form).values(**to_save))
        await session.commit()

    
    # Формирование текста анкеты
    text = f"""
    <b>Новая анкета:</b>
    📝 <b>ФИО:</b> {data['full_name']}
    🎂 <b>Возраст:</b> {data['age']}
    🌐 <b>Часовой пояс:</b> {data['timezone']}
    🏙️ <b>Город:</b> {data.get('city','Не указан')}
    📞 <b>Telegram:</b> {data.get('tg_username', 'Не указан')}
    📧 <b>Альтернативный контакт:</b> {data.get('alt_contact', 'Не указан')}
    📊 <b>Уровень:</b> {data['level']}
    🔄 <b>Опыт участия:</b> {data['participated_before']}
    🎯 <b>Приоритет:</b> {data['priority']}
    🎓 <b>Цели изучения:</b> {data.get('goals','Не указаны')}
    🎨 <b>Хобби:</b> {data['hobbies']}
    💬 <b>Темы:</b> {data['topics']}
    📅 <b>Дни:</b> {', '.join(data['days'])}
    ⏰ <b>Время:</b> {data['time_of_day']}
    ⚡ <b>Строгость расписания:</b> {data['schedule_strictness']}
    💡 <b>Инициатива:</b> {data['initiative']}
    👫 <b>Предпочтительный партнер:</b> {data['preferred_partner_gender']}
    📋 <b>Заметки:</b> {data['other_notes']}
    ✅ <b>Принял правила:</b> {'Да' if data['accepted_rules'] else 'Нет'}
    ✅ <b>Принял оферту:</b> {'Да' if data['accepted_offer'] else 'Нет'}
    ✅ <b>Принял NDA:</b> {'Да' if data['accepted_nda'] else 'Нет'}
    """

    # Отправка в чат
    chat_id = settings.BOT1_CHAT_ID  # Добавьте эту переменную в config.py
    await bot.send_photo(
        chat_id=chat_id,
        photo=file_id,
        caption=text,
        parse_mode=ParseMode.HTML
    )
    await state.clear()
    await m.answer("Спасибо! Анкета сохранена. 👍", reply_markup=kb(["Другое"], row=1))

@router.message(F.text)
async def fallback(m: Message):
    await m.answer("Пожалуйста, следуйте шагам анкеты. Используйте предложенные кнопки там, где они есть.")

async def main():
    await init_db(models)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())